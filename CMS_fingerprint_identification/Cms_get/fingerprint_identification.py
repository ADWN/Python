import gevent
import grequests
from gevent import monkey;monkey.patch_socket()
import builtwith
import hashlib
import openpyxl,os
import random,re,requests
import sys,socket
import time
import urllib
from scapy.all import *
from bs4 import BeautifulSoup as BS


class fingerprint_identification(object):
    def __init__(self,domain_list,headers_list,proxy_list,time_out,association_number):
        self.domain_list        = domain_list            #域名列表
        self.headers_list       = headers_list          #headers头部列表
        self.proxy_list         = proxy_list           #proxy代理列表
        self.time_out           = time_out            #url连接超时时间        
        self.association_number = association_number #协程数
        self.ip_dict            = {}                #获取IP信息
        self.os_dict            = {}               #获取操作系统信息
        self.title_dict         = {}              #获取Title信息
        self.server_dict        = {}             #获取Server信息
        self.dev_language_dict  = {}            #获取开发语言信息
        self.cms_dict           = {}           #获取CMS信息                                
        self.js_frame_dict      = {}          #获取JS框架信息
        self.banner_header      = {}         #获取Banner指纹库识别中的header
        self.banner_text        = {}        #获取Banner指纹库识别中的text
        self.banner_title       = {}       #获取Banner指纹库识别中的title
        self.banner_database_finger = {}  #获取Banner指纹库指纹
        self.banner_database_finger_get()#获取Banner指纹库指纹函数
        self.banner_scan_dict   = {}   #获取Banner指纹库识别运行结果
        self.cms_scan_dict      = {}  #获取Cms指纹库识别运行结果
        self.update_hit_list    = [] #收集已识别cms指纹的hit值在指纹库中的行数及hit值，方便hit值的更新[[row,hit]]        

        
    def banner_database_finger_get(self):
        '''获取banner.xlsx文件中的指纹'''
        try :
            workbook  = openpyxl.load_workbook((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'banner.xlsx')
            worksheet = workbook.active#获取表的banner表单(默认最活跃的即是第一个表)            
            for cell1,cell2 in zip(worksheet['A'][1:],worksheet['B'][1:]): 
                self.banner_database_finger.update({cell1.value:cell2.value}) #{name:key,name:key,...}
            workbook.close()
        except : #打开文件失败则添加Not_found
            self.banner_database_finger.update({'Not_found':'Not_found'})

        
    def get_ip_os(self,domain):
        '''获取IP,操作系统类型'''
        try :
            self.ip_dict[domain] = socket.gethostbyname(domain)
        except:
            self.ip_dict[domain] = 'Not_found'
        
        try :
            response = sr1(IP(dst = domain) / ICMP(), timeout = 1, verbose = 0) #向目标域名发送ICMP包（封装在IP层）,超时1s,不显示详细信息
            if (response == None):
                self.os_dict[domain] = 'Not_found'
            elif (response[IP].ttl <= 32):
                self.os_dict[domain] = 'WINDOWS'
            elif (32 < response[IP].ttl <= 64):
                self.os_dict[domain] = 'LINUX'
            elif (64 < response[IP].ttl <= 128):
                self.os_dict[domain] = 'WIN2K/NT'
            else:
                self.os_dict[domain] = 'UNIX'
        except :
            self.os_dict[domain] = 'Not_found'

    
    def get_title_server_dev(self,url):
        '''获取Title,Server,Dev_language'''
        request = urllib.request.Request(url = url,headers = random.choice(self.headers_list)) #测试使用列表第一个url
        opener  = urllib.request.build_opener(random.choice(self.proxy_list)) #是否使用代理
        urllib.request.install_opener(opener)
        try :
            response = urllib.request.urlopen(request,timeout = self.time_out) #超时初始设置3 
            text     = response.read().decode("utf-8") 
            try :
                title = BS(text,'lxml').title.text.strip() #获取title信息，从获取的text中提取title
                self.title_dict[url] = title
            except :
                self.title_dict[url] = 'Not_found'
                
            try:
                if (response.getheader(name = "Server")): #获取server信息，如果不存在(False)返回的是“None Type”
                    server = response.getheader(name = "Server") #存在(True)则添加Server/server的值
                    self.server_dict[url] = server
                elif (response.getheader(name = "server")):
                    server = response.getheader(name = "server") 
                    self.server_dict[url] = server
                else : 
                    self.server_dict[url] = 'Not_found'
            except :
                self.server_dict[url] = 'Not_found'
            
            try:
                x_powered_by_get = response.getheader(name = "X-Powered-By") #获取开发语言信息，从获取的headers中匹配查找
                cookie_get       = response.getheader(name = "Set-Cookie")
                r_php = re.compile(r'<a[^>]*?href=(\'|")[^http][^>]*?\.php(\?|\#|\1)')
                r_jsp = re.compile(r'<a[^>]*?href=(\'|")[^http][^>]*?\.jsp(\?|\#|\1)')
                r_asp = re.compile(r'<a[^>]*?href=(\'|")[^http][^>]*?\.asp(\?|\#|\1)')
                r_aspx1 = re.compile(r'<input[^>]+name=\"__VIEWSTATE')
                r_aspx2 = re.compile(r'<a[^>]*?href=(\'|")[^http][^>]*?\.aspx(\?|\#|\1)')
                if (x_powered_by_get): #存在
                    self.dev_language_dict[url] = x_powered_by_get
                elif ((cookie_get == 'PHPSSIONID') or (re.search(r_php,text) != 'None')): #从获取的headears中的cookie_get以及text中匹配查找
                    self.dev_language_dict[url] = 'PHP'
                elif ((cookie_get == 'ASPSESSION') or (re.search(r_asp,text) != 'None')):
                    self.dev_language_dict[url] = 'ASP'
                elif ((cookie_get == 'JSESSIONID') or (re.search(r_jsp,text) != 'None')):
                    self.dev_language_dict[url] = 'JSP'
                elif ((cookie_get == 'ASP.NET_SessionId') or (response.getheader(name = "X-AspNet-Version")) or (re.search(r_aspx1,text) != 'None') or (re.search(r_aspx2,text) != 'None')):
                    self.dev_language_dict[url] = 'ASPX'  
                else:
                    self.dev_language_dict[url] = 'Not_found'
            except :
                self.dev_language_dict[url] = 'Not_found'
                
        except : #url请求失败字典值都是Not_found
            self.title_dict[url]        = 'Not_found'
            self.server_dict[url]       = 'Not_found'
            self.dev_language_dict[url] = 'Not_found'
    
    
    def get_cms_js_frame(self,url):
        try:
            dict_value = builtwith.parse(url = url)
            if ('cms' in dict_value): #cms在字典中，获取cms值
                cms_name = dict_value.get('cms')
                cms_name = str(cms_name).replace('[','').replace(']','').replace("'",'')
                self.cms_dict[url] = cms_name
            else :
                self.cms_dict[url] = 'Not_found'
                
            if ('javascript-frameworks' in dict_value): # javascript-frameworks在字典中，获取JS框架值
                js_frame = dict_value.get('javascript-frameworks')
                js_frame = str(js_frame).replace('[','').replace(']','').replace("'",'')
                self.js_frame_dict[url] = js_frame
            else :
                self.js_frame_dict[url] = 'Not_found'
        except :
            self.cms_dict[url] = 'Not_found'
            self.js_frame_dict[url] = 'Not_found'   
    
    
    def banner_web_data_get(self,url):
        '''获取url列表及返回响应包的title,headers,text'''
        request = urllib.request.Request(url = url,headers = random.choice(self.headers_list)) 
        opener = urllib.request.build_opener(random.choice(self.proxy_list)) #是否使用代理
        urllib.request.install_opener(opener)
        try :
            response = urllib.request.urlopen(request,timeout = self.time_out) #超时初始设置3 
            text     = response.read().decode("utf-8") 
            try :
                title = BS(text,'lxml').title.text.strip()
                self.banner_title[url]  = title.rstrip('\n')
                self.banner_header[url] = str(response.headers)
                self.banner_text[url]   =text
            except:
                self.banner_header[url] = str(response.headers)
                self.banner_text[url]   =text                
        except : pass    
        
    def banner_check_finger(self,title,header,body,key):#检查获取到的指纹是否在数据库的指纹中(单个指纹检测)body = text
        """指纹匹配"""
        r_title   = re.compile(r'title="(.*)"')#预编译匹配title指纹的内容
        r_header  = re.compile(r'header="(.*)"')#预编译匹配header指纹的内容
        r_body    = re.compile(r'body="(.*)"')#预编译匹配body指纹的内容            
        try : #如果指纹在指定的字符串中则返回True
            if ('title="' in key):
                if re.findall(r_title, key)[0].lower() in title.lower():
                    return True
            elif ('body="' in key):
                if re.findall(r_body, key)[0] in body: 
                    return True
            else : #除了title和body其他的都在headers中进行匹配
                if (re.findall(r_header, key)[0] in header): 
                    return True
        except : pass     
    
    def banner_match_finger(self,url,title,header,body,name,key):
        '''从指纹库提取指纹，并将多个逻辑符连接一起的指纹分割成单一的指纹，并判断'''
        r_and = re.compile(r'.*\(.+\|\|.+\).*')  #预编译匹配并下有与指纹
        r_or  = re.compile(r'.*\(.+\\&.+\).*')  #预编译匹配与下有并指纹
        if ('||' in key and '&&' not in key):  #只有与'||','||.||','||.||.||'
            for rule in key.split('||'):
                if (self.banner_check_finger(title,header,body,rule)):
                    self.banner_scan_dict[url].append(name) #同一'url'键的值列表添加不同'key'值
                    break
        elif ('&&' in key and '||' not in key): #只有并'&&','&&.&&','&&.&&.&&'
            num = 0
            for rule in key.split('&&'): #分割key
                if (self.banner_check_finger(title,header,body,rule)):
                    num += 1
            if (num == len(key.split('&&'))): #key所有字符都匹配成功，添加指纹名
                self.banner_scan_dict[url].append(name)    
        elif ('||' not in key and '&&' not in key ): #与，并都没有
            if (self.banner_check_finger(title,header,body,key)):
                self.banner_scan_dict[url].append(name) 
        elif (re.findall(r_and,key)) : #并下有与
            for rule1 in key.split('&&'):
                num = 0
                if ('||' in rule1):
                    for rule2 in rule1.split('||'):
                        if (self.banner_check_finger(title,header,body,rule2)):
                            num += 1
                            break
                else:
                    if (self.banner_check_finger(title,header,body,rule1)):
                        num += 1
            if (num == len(key.split('&&'))):#条件都满足才添加指纹
                self.banner_scan_dict[url].append(name) 
        elif (re.findall(r_or,key)): #与下有并
            for rule1 in key.split('||'):
                if ('&&' in rule1):
                    num = 0
                    for rule2 in rule1.split('&&'):
                        if (self.banner_check_finger(title,header,body,rule2)):
                            num += 1
                    if (num == len(rule1.split('&&'))):#条件都满足才添加指纹
    
                        self.banner_scan_dict[url].append(name) 
                        break
                else:
                    if self.banner_check_finger(title,header,body,rule1):
                        self.banner_scan_dict[url].append(name)
                        break
        else : pass
        
        
    def md5_get(self,value):
        '''得到value的MD5值'''
        x = hashlib.md5()
        x.update(value)
        md5_value = x.hexdigest()
        return md5_value
    
    def cms_get(self,url,name,path,match_pattern,options,hit,number):
        '''指纹匹配，成功则找到CMS'''      
        request = urllib.request.Request(url = url + path,headers = random.choice(self.headers_list)) 
        opener  = urllib.request.build_opener(random.choice(self.proxy_list)) #是否使用代理
        urllib.request.install_opener(opener)
        try :
            response = urllib.request.urlopen(request,timeout = 3) #超时初始设置3 
            if (options == 'keyword'):
                text    = response.read().decode('utf-8') #只获取解码后文本
                if (match_pattern in text.lower()):
                    self.cms_scan_dict[url].append([name,hit]) #添加已匹配的cms名和hit数
                    self.update_hit_list.append([number + 2,hit]) 
                else : pass
            elif (options == 'md5'):
                content   = response.read() #只获取二进制文件（用来获取MD5值）
                if (match_pattern == self.md5_get(value = content)):
                    self.cms_scan_dict[url].append([name,hit]) #添加已匹配的cms名和hit数
                    self.update_hit_list.append([number + 2,hit])
                else : pass
            else : pass
        except :  pass    
        
    
    def get_result_dict(self):
        '''获取结果列表：[value,value,...]'''
        process_list,result_list,fail_match_url_list = [],[],[] #协程池列表，结果列表，未匹配cms的url列表
        cms_list =['08cms', '1039_jxt', '1039\xe5\xae\xb6\xe6\xa0\xa1\xe9\x80\x9a', '3gmeeting', '3gmeeting\xe8\xa7\x86\xe8\xae\xaf\xe7\xb3\xbb\xe7\xbb\x9f', '51fax\xe4\xbc\xa0\xe7\x9c\x9f\xe7\xb3\xbb\xe7\xbb\x9f', '53kf', '5ucms', '686_weixin', '6kbbs', '74cms', '86cms', 'afterlogicwebmail\xe7\xb3\xbb\xe7\xbb\x9f', 'appcms', 'aspcms', 'b2bbuilder', 'beescms', 'bookingecms\xe9\x85\x92\xe5\xba\x97\xe7\xb3\xbb\xe7\xbb\x9f', 'cactiez\xe6\x8f\x92\xe4\xbb\xb6', 'chinacreator', 'cxcms', 'dk\xe5\x8a\xa8\xe7\xa7\x91cms', 'doyo\xe9\x80\x9a\xe7\x94\xa8\xe5\xbb\xba\xe7\xab\x99\xe7\xb3\xbb\xe7\xbb\x9f', 'dtcms', 'dvrdvs-webs', 'datalifeengine', 'dayucms', 'dedecms', 'destoon', 'digital campus2.0', 'digitalcampus2.0', 'discuz', 'discuz7.2', 'drupal', 'dswjcms', 'duomicms', 'dvbbs', 'dzzoffice', 'ecshop', 'ec_word\xe4\xbc\x81\xe4\xb8\x9a\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'emlog', 'easysite\xe5\x86\x85\xe5\xae\xb9\xe7\xae\xa1\xe7\x90\x86', 'edusoho', 'empirecms', 'epaper\xe6\x8a\xa5\xe5\x88\x8a\xe7\xb3\xbb\xe7\xbb\x9f', 'epoint', 'espcms', 'fengcms', 'foosuncms', 'gentlecms', 'gever', 'glassfish', 'h5\xe9\x85\x92\xe5\xba\x97\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'hdwiki', 'hjcms\xe4\xbc\x81\xe4\xb8\x9a\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'himail', 'hishop\xe5\x95\x86\xe5\x9f\x8e\xe7\xb3\xbb\xe7\xbb\x9f', 'hituxcms', 'ilas\xe5\x9b\xbe\xe4\xb9\xa6\xe7\xb3\xbb\xe7\xbb\x9f', 'iloanp2p\xe5\x80\x9f\xe8\xb4\xb7\xe7\xb3\xbb\xe7\xbb\x9f', 'imo\xe4\xba\x91\xe5\x8a\x9e\xe5\x85\xac\xe5\xae\xa4\xe7\xb3\xbb\xe7\xbb\x9f', 'insightsoft', 'iwebshop', 'iwmscms', 'jboos', 'jishigou', 'jeecms', 'jingyi', 'joomla', 'kangle\xe8\x99\x9a\xe6\x8b\x9f\xe4\xb8\xbb\xe6\x9c\xba', 'kesioncms', 'kessioncms', 'kingcms', 'lebishop\xe7\xbd\x91\xe4\xb8\x8a\xe5\x95\x86\xe5\x9f\x8e', 'live800', 'live800\xe6\x8f\x92\xe4\xbb\xb6', 'ljcms', 'mlecms', 'mailgard', 'majexpress', 'mallbuilder', 'maticsoftsns', 'minyoocms', 'mvmmall', 'mymps\xe8\x9a\x82\xe8\x9a\x81\xe5\x88\x86\xe7\xb1\xbb\xe4\xbf\xa1\xe6\x81\xaf', 'n\xe7\x82\xb9\xe8\x99\x9a\xe6\x8b\x9f\xe4\xb8\xbb\xe6\x9c\xba', 'opensns', 'ourphp', 'php168', 'phpcms', 'phpwind', 'phpok', 'piw\xe5\x86\x85\xe5\xae\xb9\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'phpmyadmin', 'phpwind\xe7\xbd\x91\xe7\xab\x99\xe7\xa8\x8b\xe5\xba\x8f', 'pigcms', 'powercreator\xe5\x9c\xa8\xe7\xba\xbf\xe6\x95\x99\xe5\xad\xa6\xe7\xb3\xbb\xe7\xbb\x9f', 'powereasy', 'sapnetweaver', 'shopex', 'shop7z', 'shopnc\xe5\x95\x86\xe5\x9f\x8e\xe7\xb3\xbb\xe7\xbb\x9f', 'shopnum', 'siteserver', 'soullon', 'southidc', 'supesite', 't-site\xe5\xbb\xba\xe7\xab\x99\xe7\xb3\xbb\xe7\xbb\x9f', 'theol\xe7\xbd\x91\xe7\xbb\x9c\xe6\x95\x99\xe5\xad\xa6\xe7\xbb\xbc\xe5\x90\x88\xe5\xb9\xb3\xe5\x8f\xb0', 'trs\xe8\xba\xab\xe4\xbb\xbd\xe8\xae\xa4\xe8\xaf\x81\xe7\xb3\xbb\xe7\xbb\x9f', 'tipask\xe9\x97\xae\xe7\xad\x94\xe7\xb3\xbb\xe7\xbb\x9f', 'tomcat', 'trsids', 'trunkey', 'turbomail\xe9\x82\xae\xe7\xae\xb1\xe7\xb3\xbb\xe7\xbb\x9f', 'v2\xe8\xa7\x86\xe9\xa2\x91\xe4\xbc\x9a\xe8\xae\xae\xe7\xb3\xbb\xe7\xbb\x9f', 'v5shop', 'venshop2010\xe5\x87\xa1\xe4\xba\xba\xe7\xbd\x91\xe7\xbb\x9c\xe8\xb4\xad\xe7\x89\xa9\xe7\xb3\xbb\xe7\xbb\x9f', 'vos3000', 'veryide', 'wcm\xe7\xb3\xbb\xe7\xbb\x9fv6', 'wordpress', 'ws2004\xe6\xa0\xa1\xe5\x9b\xad\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'wangzt', 'weblogic', 'webmail', 'weboffice', 'webnet cms', 'webnetcms', 'wilmaroa\xe7\xb3\xbb\xe7\xbb\x9f', 'winmail server', 'winmailserver', 'wizbank', 'xplus\xe6\x8a\xa5\xe7\xa4\xbe\xe7\xb3\xbb\xe7\xbb\x9f', 'xpshop', 'yidacms', 'yongyou', 'z-blog', 'zabbix', 'zoomla', 'abcms', 'able_g2s', 'acsno', 'acsoft', 'actcms', 'adtsec_gateway', 'akcms', 'anleye', 'anmai', 'anmai\xe5\xae\x89\xe8\x84\x89\xe6\x95\x99\xe5\x8a\xa1\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'anymacromail', 'apabi_tasi', 'asiastar_sm', 'aten_kvm', 'atripower', 'avcon6', 'axis2', 'ayacms', 'b2cgroup', 'baiaozhi', 'beidou', 'bluecms', 'boblog', 'bocweb', 'bohoog', 'bytevalue_router', 'canon', 'chamilo-lms', 'ckfinder', 'cmseasy', 'cmstop', 'cnoa', 'codeigniter', 'comexe_ras', 'cscms', 'cutecms', 'd-link', 'dahua_dss', 'daiqile_p2p', 'dalianqianhao', 'damall', 'damicms', 'dfe_scada', 'dianyips', 'diguocms\xe5\xb8\x9d\xe5\x9b\xbd', 'dircms', 'dkcms', 'dossm', 'douphp', 'dreamgallery', 'dubbo', 'eshangbao\xe6\x98\x93\xe5\x95\x86\xe5\xae\x9d', 'easethink', 'easy7\xe8\xa7\x86\xe9\xa2\x91\xe7\x9b\x91\xe6\x8e\xa7\xe5\xb9\xb3\xe5\x8f\xb0', 'ecweb_shop', 'edayshop', 'edjoy', 'eduplate', 'edusohocms', 'eims', 'eimscms', 'electric_monitor', 'empire_cms', 'enableq', 'enjie_soft', 'es-cloud', 'esafenet_dlp', 'esccms', 'ewebs', 'expocms', 'extmail', 'eyou', 'e\xe5\x88\x9b\xe7\xab\x99', 'fang5173', 'fangwei', 'fastmeeting', 'fcms', 'fcms\xe6\xa2\xa6\xe6\x83\xb3\xe5\xbb\xba\xe7\xab\x99', 'feifeicms', 'feiyuxing_router', 'finecms', 'fiyocms', 'foosun', 'foosun\xe6\x96\x87\xe7\xab\xa0\xe7\xb3\xbb\xe7\xbb\x9f', 'fsmcms', 'gbcom_wlan', 'genixcms', 'gnuboard', 'gocdkey', 'gooine_sqjz', 'gowinsoft_jw', 'gxcms', 'hac_gateway', 'haitianoa', 'hanweb', 'haohan', 'heeroa', 'hf_firewall', 'hongzhi', 'horde_email', 'house5', 'hsort', 'huachuang_router', 'huanet', 'huashi_tv', 'humhub', 'idvr', 'ipowercms', 'iceflow_vpn_router', 'ideacms', 'ieadcms', 'iflytek_soft', 'igenus', 'ikuai', 'insight', 'jenkins', 'jienuohan', 'jieqicms', 'jindun_gateway', 'jingci_printer', 'jinpan', 'jinqiangui_p2p', 'jishitongxun', 'joomle', 'jumbotcms', 'juniper_vpn', 'kill_firewall', 'kingdee_eas', 'kingdee_oa', 'kinggate', 'kingosoft_xsweb', 'kj65n_monitor', 'klemanndesign', 'kuwebs', 'kxmail', 'landray', 'lebishop', 'lezhixing_datacenter', 'lianbangsoft', 'liangjing', 'libsys', 'linksys', 'looyu_live', 'ltpower', 'luepacific', 'luzhucms', 'lvmaque', 'maccms', 'magento', 'mailgard-webmail', 'mainone_b2b', 'maopoa', 'maxcms', 'mbbcms', 'metinfo', 'mikrotik_router', 'moxa_nport_router', 'mpsec', 'myweb', 'nanjing_shiyou', 'natshell', 'nbcms', 'net110', 'netcore', 'netgather', 'netoray_nsg', 'netpower', 'newvane_onlineexam', 'nitc', 'nitc\xe5\xae\x9a\xe6\xb5\xb7\xe7\xa5\x9e\xe7\x9c\x9f', 'niubicms', 'ns-asg', 'otcms', 'pageadmin', 'panabit', 'phpb2b', 'phpcmsv9', 'phpdisk', 'phpmaps', 'phpmps', 'phpmywind', 'phpshe', 'phpshop', 'phpvibe', 'phpweb', 'phpwiki', 'phpyun', 'piaoyou', 'pkpmbs', 'plc_router', 'powercreator', 'qht_study', 'qianbocms', 'qibosoft', 'qiuxue', 'qizhitong_manager', 'qzdatasoft\xe5\xbc\xba\xe6\x99\xba\xe6\x95\x99\xe5\x8a\xa1\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', 'rockoa', 'rockontrol', 'ruijie_router', 'ruvar_oa', 'ruvarhrm', 's8000', 'santang', 'sdcms', 'seagate_nas', 'seawind', 'seentech_uccenter', 'sgc8000', 'shadows-it', 'shenlan_jiandu', 'shlcms', 'shopnum1', 'shopxp', 'shuangyang_oa', 'siteengine', 'sitefactory', 'skypost', 'skytech', 'smart_oa', 'soffice', 'soullon_edu', 'srun_gateway', 'star-net', 'startbbs', 'strongsoft', 'subeicms', 'syncthru_web_service', 'synjones_school', 'syxxjs', 'sztaiji_zw', 'taocms', 'taodi', 'terramaster', 'thinkox', 'thinkphp', 'thinksns', 'tianbo_train', 'tianrui_lib', 'tipask', 'tongdaoa', 'topsec', 'totalsoft_lib', 'tp-link', 'trs_ids', 'trs_inforadar', 'trs_lunwen', 'trs_wcm', 'typecho', 'umail', 'uniflows', 'unis_gateway', 'uniwin_gov', 'urp', 'v2_conference', 'vbulletin', 'vicworl', 'visionsoft_velcro', 'wangqushop', 'wdcp', 'wdscms', 'weaver_oa', 'websitebaker', 'wecenter', 'weixinpl', 'weway_soft', 'wisedu_elcs', 'workyisystem', 'workyi_system', 'wygxcms', 'xdcms', 'xiaowuyou_cms', 'xikecms', 'xinhaisoft', 'xinyang', 'xinzuobiao', 'xplus', 'xr_gatewayplatform', 'xuezi_ceping', 'xycms', 'ynedut_campus', 'yongyou_a8', 'yongyou_crm', 'yongyou_ehr', 'yongyou_fe', 'yongyou_icc', 'yongyou_nc', 'yongyou_u8', 'yongyou_zhiyuan_a6', 'yuanwei_gateway', 'yxlink', 'zblog', 'zcncms', 'zdsoft_cnet', 'zentao', 'zeroboard', 'zf_cms', 'zfsoft', 'zhongdongli_school', 'zhonghaida_vnet', 'zhongqidonglicms', 'zhongruan_firewall', 'zhoupu', 'zhuangxiu', 'zhuhaigaoling_huanjingzaosheng', 'zmcms', 'zmcms\xe5\xbb\xba\xe7\xab\x99', 'zte', 'zuitu', 'zzcms', '\xe4\xb8\x87\xe4\xbc\x97\xe7\x94\xb5\xe5\xad\x90\xe6\x9c\x9f\xe5\x88\x8acms', '\xe4\xb8\x87\xe5\x8d\x9a\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f2006', '\xe4\xb8\x87\xe5\x8d\x9a\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe4\xb8\x87\xe6\x88\xb7oa', '\xe4\xb8\x87\xe6\xac\xa3\xe9\xab\x98\xe6\xa0\xa1\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe4\xb8\x89\xe6\x89\x8d\xe6\x9c\x9f\xe5\x88\x8a\xe7\xb3\xbb\xe7\xbb\x9f', '\xe4\xb8\xad\xe4\xbc\x81\xe5\x8a\xa8\xe5\x8a\x9bcms', '\xe4\xb9\x90\xe5\xbd\xbc\xe5\xa4\x9a\xe7\xbd\x91\xe5\xba\x97', '\xe4\xba\xbf\xe9\x82\xaeemail', '\xe4\xbc\x81\xe6\x99\xba\xe9\x80\x9a\xe7\xb3\xbb\xe5\x88\x97\xe4\xb8\x8a\xe7\xbd\x91\xe8\xa1\x8c\xe4\xb8\xba\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe4\xbc\x97\xe6\x8b\x93', '\xe5\x85\xa8\xe7\xa8\x8boa', '\xe5\x87\xa1\xe8\xaf\xba\xe4\xbc\x81\xe4\xb8\x9a\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\x88\x86\xe7\xb1\xbb\xe4\xbf\xa1\xe6\x81\xaf\xe7\xbd\x91bank.asp\xe5\x90\x8e\xe9\x97\xa8', '\xe5\x88\x9b\xe6\x8d\xb7\xe9\xa9\xbe\xe6\xa0\xa1\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\x8d\x8e\xe5\xa4\x8f\xe5\x88\x9b\xe6\x96\xb0appex\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\x8d\x97\xe6\x96\xb9\xe6\x95\xb0\xe6\x8d\xae', '\xe5\x8f\xa3\xe7\xa6\x8f\xe7\xa7\x91\xe6\x8a\x80', '\xe5\x91\xb3\xe5\xa4\x9a\xe7\xbe\x8e\xe5\xaf\xbc\xe8\x88\xaa', '\xe5\x95\x86\xe5\xa5\x87cms', '\xe5\x95\x86\xe5\xae\xb6\xe4\xbf\xa1\xe6\x81\xaf\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\x9b\x9b\xe9\x80\x9a\xe6\x94\xbf\xe5\xba\x9c\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xa4\xa7\xe6\xb1\x89jcms', '\xe5\xa4\xa9\xe6\x9f\x8f\xe5\x9c\xa8\xe7\xba\xbf\xe8\x80\x83\xe8\xaf\x95\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xa4\xa9\xe8\x9e\x8d\xe4\xbf\xa1panabit', '\xe5\xae\x81\xe5\xbf\x97\xe5\xad\xa6\xe6\xa0\xa1\xe7\xbd\x91\xe7\xab\x99', '\xe5\xae\x81\xe5\xbf\x97\xe5\xad\xa6\xe6\xa0\xa1\xe7\xbd\x91\xe7\xab\x99\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xae\x89\xe4\xb9\x90\xe4\xb8\x9a\xe6\x88\xbf\xe4\xba\xa7\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xae\x9a\xe6\xb5\xb7\xe7\xa5\x9e\xe7\x9c\x9f', '\xe5\xb0\x8f\xe8\xae\xa1\xe5\xa4\xa9\xe7\xa9\xba\xe8\xbf\x9b\xe9\x94\x80\xe5\xad\x98\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xb0\x98\xe6\x9c\x88\xe4\xbc\x81\xe4\xb8\x9a\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xb0\x98\xe7\xbc\x98\xe9\x9b\x85\xe5\xa2\x83\xe5\x9b\xbe\xe6\x96\x87\xe7\xb3\xbb\xe7\xbb\x9f', '\xe5\xbb\xba\xe7\xab\x99\xe4\xb9\x8b\xe6\x98\x9f', '\xe5\xbe\xae\xe6\x93\x8e\xe7\xa7\x91\xe6\x8a\x80', '\xe6\x82\x9f\xe7\xa9\xbacrm', '\xe6\x82\x9f\xe7\xa9\xbacrm\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x93\x8e\xe5\xa4\xa9\xe6\x94\xbf\xe5\x8a\xa1\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x96\xb0\xe4\xb8\xba\xe8\xbd\xaf\xe4\xbb\xb6e-learning\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x96\xb0\xe7\xa7\x80', '\xe6\x96\xb9\xe7\xbb\xb4\xe5\x9b\xa2\xe8\xb4\xad', '\xe6\x96\xb9\xe7\xbb\xb4\xe5\x9b\xa2\xe8\xb4\xad\xe8\xb4\xad\xe7\x89\xa9\xe5\x88\x86\xe4\xba\xab\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x97\xb6\xe4\xbb\xa3\xe4\xbc\x81\xe4\xb8\x9a\xe9\x82\xae', '\xe6\x98\x8e\xe8\x85\xbecms', '\xe6\x98\x93\xe5\x88\x9b\xe6\x80\x9d', '\xe6\x98\x93\xe5\x88\x9b\xe6\x80\x9d\xe6\x95\x99\xe8\x82\xb2\xe5\xbb\xba\xe7\xab\x99\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x98\x93\xe6\x83\xb3cms', '\xe6\x99\xba\xe7\x9d\xbf\xe7\xbd\x91\xe7\xab\x99\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x9c\x80\xe5\x9c\x9f\xe5\x9b\xa2\xe8\xb4\xad\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x9c\xaa\xe7\x9f\xa5oem\xe5\xae\x89\xe9\x98\xb2\xe7\x9b\x91\xe6\x8e\xa7\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x9c\xaa\xe7\x9f\xa5\xe6\x94\xbf\xe5\xba\x9c\xe9\x87\x87\xe8\xb4\xad\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x9c\xaa\xe7\x9f\xa5\xe6\x9f\xa5\xe8\xaf\xa2\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\x9d\xad\xe5\xb7\x9e\xe5\x8d\x9a\xe9\x87\x87cms', '\xe6\x9d\xb0\xe5\xa5\x87\xe5\xb0\x8f\xe8\xaf\xb4\xe8\xbf\x9e\xe8\xbd\xbd\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\xa1\x83\xe6\xba\x90\xe7\x9b\xb8\xe5\x86\x8c\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\xb1\x87\xe6\x88\x90\xe4\xbc\x81\xe4\xb8\x9a\xe5\xbb\xba\xe7\xab\x99cms', '\xe6\xb1\x87\xe6\x96\x87\xe5\x9b\xbe\xe4\xb9\xa6\xe9\xa6\x86\xe4\xb9\xa6\xe7\x9b\xae\xe6\xa3\x80\xe7\xb4\xa2\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\xb1\x89\xe7\xa0\x81\xe9\xab\x98\xe6\xa0\xa1\xe6\xaf\x95\xe4\xb8\x9a\xe7\x94\x9f\xe5\xb0\xb1\xe4\xb8\x9a\xe4\xbf\xa1\xe6\x81\xaf\xe7\xb3\xbb\xe7\xbb\x9f', '\xe6\xb3\x9b\xe5\xbe\xaee-office', '\xe6\xb3\x9b\xe5\xbe\xaeoa', '\xe6\xb5\xaa\xe6\xbd\xaecms', '\xe6\xb5\xb7\xe5\xba\xb7\xe5\xa8\x81\xe8\xa7\x86', '\xe7\x88\xb1\xe6\xb7\x98\xe5\xae\xa2', '\xe7\x88\xb1\xe8\xa3\x85\xe7\xbd\x91', '\xe7\x94\xa8\xe5\x8f\x8bfe\xe5\x8d\x8f\xe4\xbd\x9c\xe5\x8a\x9e\xe5\x85\xac\xe5\xb9\xb3\xe5\x8f\xb0', '\xe7\x94\xa8\xe5\x8f\x8bfe\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe7\x94\xa8\xe5\x8f\x8bturbcrm\xe7\xb3\xbb\xe7\xbb\x9f', '\xe7\x94\xa8\xe5\x8f\x8bu8', '\xe7\x94\xa8\xe5\x8f\x8b', '\xe7\x9a\x93\xe7\xbf\xb0\xe9\x80\x9a\xe7\x94\xa8\xe6\x95\xb0\xe5\xad\x97\xe5\x8c\x96\xe6\xa0\xa1\xe5\x9b\xad\xe5\xb9\xb3\xe5\x8f\xb0', '\xe7\x9c\x81\xe7\xba\xa7\xe5\x86\x9c\xe6\x9c\xba\xe6\x9e\x84\xe7\xbd\xae\xe8\xa1\xa5\xe8\xb4\xb4\xe4\xbf\xa1\xe6\x81\xaf\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe7\xa7\x91\xe4\xbf\xa1\xe9\x82\xae\xe4\xbb\xb6\xe7\xb3\xbb\xe7\xbb\x9f', '\xe7\xa7\x91\xe8\xbf\x88ras', '\xe7\xa8\x8b\xe6\xb0\x8f\xe8\x88\x9e\xe6\x9b\xb2cms', '\xe7\xbb\xbf\xe9\xba\xbb\xe9\x9b\x80\xe5\x80\x9f\xe8\xb4\xb7\xe7\xb3\xbb\xe7\xbb\x9f', '\xe7\xbd\x91\xe8\xb6\xa3\xe5\x95\x86\xe5\x9f\x8e', '\xe7\xbd\x91\xe9\x92\x9b\xe6\x96\x87\xe7\xab\xa0\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe8\x80\x81y\xe6\x96\x87\xe7\xab\xa0\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe8\x81\x94\xe4\xbc\x97mediinfo\xe5\x8c\xbb\xe9\x99\xa2\xe7\xbb\xbc\xe5\x90\x88\xe7\xae\xa1\xe7\x90\x86\xe5\xb9\xb3\xe5\x8f\xb0', '\xe8\x87\xaa\xe5\x8a\xa8\xe5\x8f\x91\xe5\x8d\xa1\xe5\xb9\xb3\xe5\x8f\xb0', '\xe8\x89\xaf\xe7\xb2\xbe\xe5\x8d\x97\xe6\x96\xb9', '\xe8\x89\xba\xe5\xb8\x86cms', '\xe8\x8f\xb2\xe6\x96\xaf\xe7\x89\xb9\xe8\xaf\xba\xe6\x9c\x9f\xe5\x88\x8a\xe7\xb3\xbb\xe7\xbb\x9f', '\xe8\x93\x9d\xe5\x87\x8ceis\xe6\x99\xba\xe6\x85\xa7\xe5\x8d\x8f\xe5\x90\x8c\xe5\xb9\xb3\xe5\x8f\xb0', '\xe8\x93\x9d\xe7\xa7\x91cms', '\xe8\x96\x84\xe5\x86\xb0\xe6\x97\xb6\xe6\x9c\x9f\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe8\xae\xaf\xe6\x97\xb6\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9fcms', '\xe8\xae\xb0\xe4\xba\x8b\xe7\x8b\x97', '\xe8\xb4\xb7\xe9\xbd\x90\xe4\xb9\x90\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x80\x9a\xe8\xbe\xbeoa\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x80\x9f\xe8\xb4\x9dcms', '\xe9\x87\x91\xe8\x89\xb2\xe6\xa0\xa1\xe5\x9b\xad', '\xe9\x87\x91\xe8\x9d\xb6oa', '\xe9\x87\x91\xe8\x9d\xb6\xe5\x8d\x8f\xe4\xbd\x9c\xe5\x8a\x9e\xe5\x85\xac\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x87\x91\xe9\x92\xb1\xe6\x9f\x9cp2p', '\xe9\x9b\x86\xe6\x97\xb6\xe9\x80\x9a\xe8\xae\xaf\xe7\xa8\x8b\xe5\xba\x8f', '\xe9\x9c\xb2\xe7\x8f\xa0\xe6\x96\x87\xe7\xab\xa0\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x9d\x92\xe4\xba\x91\xe5\xae\xa2cms', '\xe9\x9d\x92\xe5\xb3\xb0\xe7\xbd\x91\xe7\xbb\x9c\xe6\x99\xba\xe8\x83\xbd\xe7\xbd\x91\xe7\xab\x99\xe7\xae\xa1\xe7\x90\x86\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x9d\x92\xe6\x9e\x9c\xe5\xad\xa6\xe7\x94\x9f\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x9d\x92\xe6\x9e\x9c\xe5\xad\xa6\xe7\x94\x9f\xe7\xbb\xbc\xe5\x90\x88\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x9d\x92\xe6\x9e\x9c\xe6\x95\x99\xe5\x8a\xa1\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x9d\x92\xe6\x9e\x9c\xe8\xbd\xaf\xe4\xbb\xb6\xe6\x95\x99\xe5\x8a\xa1\xe7\xb3\xbb\xe7\xbb\x9f', '\xe9\x9d\x9e\xe5\x87\xa1\xe5\xbb\xba\xe7\xab\x99']        
        pool= gevent.pool.Pool(self.association_number)  
        for domain in self.domain_list:
            self.ip_dict.update({domain:''})
            self.os_dict.update({domain:''})             
            self.title_dict.update({'http://' + domain:''})
            self.server_dict.update({'http://' + domain:''})
            self.dev_language_dict.update({'http://' + domain:''})  
            process_list.append(pool.spawn(self.get_ip_os, domain))
            process_list.append(pool.spawn(self.get_title_server_dev,'http://' + domain)) 
        gevent.joinall(process_list)     
        process_list = []                                #清空协程池
        for domain in self.domain_list:   
            self.cms_dict.update({'http://' + domain:''}) #初始化
            self.js_frame_dict.update({'http://' + domain:''})
            process_list.append(pool.spawn(self.get_cms_js_frame,'http://' + domain))
        gevent.joinall(process_list)  
        for domain in self.domain_list:
            self.banner_title.update({'http://' + domain:'Not_found'})  #初始化
            self.banner_header.update({'http://' + domain:'Not_found'}) #初始化
            self.banner_text.update({'http://' + domain:'Not_found'})   #初始化
            process_list.append(pool.spawn(self.banner_web_data_get,'http://' + domain))
        gevent.joinall(process_list) 
        process_list = []                                #清空协程池
        for url in self.banner_title:
            self.banner_scan_dict.update({url:[]})       #初始化self.banner_scan_dict字典
            for name in self.banner_database_finger:
                process_list.append(pool.spawn(self.banner_match_finger,url,self.banner_title[url],self.banner_header[url],self.banner_text[url],name,self.banner_database_finger[name])) 
        gevent.joinall(process_list)
        for url in self.banner_scan_dict:
            if (len(self.banner_scan_dict[url]) != 0): #获取到banner信息
                for i in range(len(self.banner_scan_dict[url])):
                    if (self.banner_scan_dict[url][i] in cms_list):
                        self.banner_scan_dict[url].append(banner)
                        break
                    elif ((self.banner_scan_dict[url][i] not in cms_list) and (i == len(self.banner_scan_dict[url]) - 1)):
                        self.banner_scan_dict[url].append('Not_found')
                    else : pass
            else :
                self.banner_scan_dict[url].append('Not_found')
                self.banner_scan_dict[url].append('Not_found')        
        
        for i in range(len(self.domain_list)):
            result_list.append([])
            result_list[i].append(self.domain_list[i])                                            #添加domain_list中的域名
            result_list[i].append(self.title_dict['http://' + self.domain_list[i]])               #添加title_dict中的title
            result_list[i].append(self.ip_dict[self.domain_list[i]])                              #添加ip_dict中的ip
            if (self.cms_dict['http://' + self.domain_list[i]] != 'Not_found'):
                result_list[i].append(self.cms_dict['http://' + self.domain_list[i]])             #添加cms_dict中的cms
            else :
                result_list[i].append(self.banner_scan_dict['http://' + self.domain_list[i]][-1]) #添加banner_scan_dict中的cms
            result_list[i].append(str(self.banner_scan_dict['http://' + self.domain_list[i]][:-1]).replace('[','').replace(']','').replace("'",'')) #添加banner_scan_dict中的banner信息                                            #添加Banner初始设置为Not_found
            result_list[i].append(self.os_dict[self.domain_list[i]])                              #添加os_dict中的os
            result_list[i].append(self.server_dict['http://' + self.domain_list[i]])              #添加server_dict中的server
            result_list[i].append(self.js_frame_dict['http://' + self.domain_list[i]])            #添加js_frame_dict中js_frame
            result_list[i].append(self.dev_language_dict['http://' + self.domain_list[i]])        #添加dev_langeuage_dict中的dev_langeuage        
        
        for value in result_list: 
            if (value[3] == 'Not_found'):
                fail_match_url_list.append('http://' + value[0])
            else : pass      
        #for domain in self.domain_list:
            #fail_match_url_list.append('http://' + domain)

        if (len(fail_match_url_list) != 0): #存在没有匹配出cms的域名网站
            cms_finger_name_get,cms_finger_path_get,cms_finger_match_pattern_get,cms_finger_options_get,cms_finger_hit_get = [],[],[],[],[]
            workbook  = openpyxl.load_workbook((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'cms.xlsx')
            worksheet = workbook.active#获取表的banner表单(默认最活跃的即是第一个表)            
            for cell1,cell2,cell3,cell4,cell5 in zip(worksheet['A'][1:],worksheet['B'][1:],worksheet['C'][1:],worksheet['D'][1:],worksheet['E'][1:]):
                cms_finger_name_get.append(cell1.value)
                cms_finger_path_get.append(cell2.value)
                cms_finger_match_pattern_get.append(cell3.value)
                cms_finger_options_get.append(cell4.value)
                cms_finger_hit_get.append(cell5.value)
            workbook.close()
            process_list = []
            pool= gevent.pool.Pool(self.association_number)  
            for url in fail_match_url_list:
                self.cms_scan_dict.update({url:[]})  #初始化self.cms_scan_dict字典,{'url':[],...}     
                for i in range(len(cms_finger_name_get)):
                    process_list.append(pool.spawn(self.cms_get,url,cms_finger_name_get[i],cms_finger_path_get[i],cms_finger_match_pattern_get[i],cms_finger_options_get[i],cms_finger_hit_get[i],i))
            gevent.joinall(process_list)
            for url in self.cms_scan_dict:
                if (len(self.cms_scan_dict[url]) == 0):   #没有找到cms
                    self.cms_scan_dict[url].append(['Not_found']) 
                elif (len(self.cms_scan_dict[url]) == 1): #匹配到的cms值只有一个，去除列表中的hit值  
                    del self.cms_scan_dict[url][0][-1]
                else:                                     #匹配到的cms值不只有一个，比较hit值大小，取hit值较大的cms
                    max_hit  = 0                          #获取最大的hit值
                    cms_list = []                         #获取最大的hit值对应的cms名
                    for i in range(len(self.cms_scan_dict[url])):
                        if (max_hit < self.cms_scan_dict[url][i][1]):
                            max_hit = self.cms_scan_dict[url][i][1]
                        else : pass
    
                    for i in range(len(self.cms_scan_dict[url])):
                        if (self.cms_scan_dict[url][i][1] == max_hit):
                            cms_list.append(self.cms_scan_dict[url][i][0])
                        else : pass
                    self.cms_scan_dict.update({url:[]}) #初始化    
                    self.cms_scan_dict[url].append(cms_list)
            print(self.cms_scan_dict)
                
            #if (len(self.update_hit_list) != 0): #获取到了已匹配到的cms的hit值，更新指纹库中的对应指纹的hit值
                #workbook  = openpyxl.load_workbook((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'cms.xlsx')
                #worksheet = workbook.active#获取表的banner表单(默认最活跃的即是第一个表)
                #for i in range(len(self.update_hit_list)):
                    #worksheet.cell(row = self.update_hit_list[i][0],column = 5,value = self.update_hit_list[i][1] + 1)
                #workbook.save((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'cms.xlsx')
                #workbook.close        
            #else : pass  
            
        for value in result_list:
            for url in self.cms_scan_dict:
                if(value[0] == url.replace('http://','')):#添加到结果输出列表中
                    value[3] = str(self.cms_scan_dict[url]).replace('[','').replace(']','').replace("'",'') #添加cms信息
                else : pass            
        else :pass
        
        return result_list
         
  

