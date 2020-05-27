import gevent
import grequests
from gevent import monkey;monkey.patch_socket()
import codecs,csv
import hashlib
import openpyxl,os
import random,re,requests
import sys
import time
import urllib
import xlwt

class user_function(object):
    def __init__(self):
        self.proxy_url   = 'https://www.xicidaili.com/nn/'  
        self.proxy_file1 = (os.getcwd() + '/Cms_get/').replace('\\','/') + 'proxies_from_xiciproxy.txt'
        self.proxy_file2 = (os.getcwd() + '/Cms_get/').replace('\\','/') + 'proxies.txt'
        self.proxy_url_list       = [] #获取代理网站代理IP
        self.effective_proxy_list = [] #获取有效代理IP列表  
        self.headers_list         = []
        self.get_headers_list()
        self.file_link_dict       = {}
        self.file_md5_list        = []        
     
    def get_headers_list(self):
        '''获取headers_list列表：[value,value,...]'''
        try :
            file_name = (os.getcwd() + '/Cms_get/').replace('\\','/') + 'User-Agents.txt'
            read_file = open(file_name,'r',encoding = 'utf-8')
            user_agents_get = read_file.readlines()
            read_file.close 
            for value in user_agents_get:
                headers = {
                    "Accept":'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    "User-Agent":value.rstrip('\n'),
                    "Upgrade-Insecure-Requests":'1',
                    "Connection":'keep-alive',
                    "Cache-Control":'max-age=0',
                    "Accept-Language":'zh-CN,zh;q=0.8',
                    "Referer": "http://www.baidu.com/link?url=www.so.com&url=www.soso.com&&url=www.sogou.com"          
                }   
                self.headers_list.append(headers)   
        except : #打开文件失败时
            self.headers_list.append({"User-Agent":'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0'})    
        
    def get_page_proxy(self,page,time_out):
        request = urllib.request.Request(url = self.proxy_url + str(page),headers = random.choice(self.headers_list))
        try :
            response = urllib.request.urlopen(request,timeout = time_out)
            text = response.read().decode('utf-8')
            proxy_ip = re.findall('<td>\d*\.\d*\.\d*\.\d*</td>',text) 
            proxy_port = re.findall('<td>\d*</td>',text) 
            for i in range(len(proxy_ip)): 
                proxy_ip[i]   = proxy_ip[i].replace('<td>','').replace('</td>','')    #获取代理IP
                proxy_port[i] = proxy_port[i].replace('<td>','').replace('</td>','') #获取代理IP端口
                self.proxy_url_list.append("{'http':'" + "http://{}:{}'".format(proxy_ip[i],proxy_port[i]) + ",'https':'"+"https://{}:{}'".format(proxy_ip[i],proxy_port[i]) + '}')
        except :                   #获取不到代理，添加默认代理IP               
            self.proxy_url_list.append("{'http':'http://211.149.252.155:8    888','https':'https://211.149.252.155:8888'}")
    
    def get_proxy_list(self,page,time_out,association_number):
        if (page.isdigit()): #输入字符串为纯数字
            if (0 < int(page) < 4055):#输入为数字且在网站页数范围内        
                process_list = []   
                pool = gevent.pool.Pool(association_number)         
                for i in range(1,int(page) + 1):
                    process_list.append(pool.spawn(self.get_page_proxy, i,time_out))
                gevent.joinall(process_list)
                proxy_ip_list = list(set(self.proxy_url_list)) #去重
                write_file = open(self.proxy_file1,'w+',encoding = 'utf-8')
                for value in proxy_ip_list:
                    write_file.write(value + '\n')
                write_file.close()   
                page_flag = 'True'      #成功获取代理
            else :
                page_flag = 'False2' #输入页数越界!\n输入页数为0或者超出代理网页总页数!
        else :
            page_flag = 'False1' #输入格式错误!\n输入页数不为纯数字!
    
        return page_flag 
    
    def check(self,time_out,proxy): 
        try :
            response = requests.get(url='http://www.baidu.com',headers = random.choice(self.headers_list),timeout = time_out,proxies = proxy)
            if (response.status_code == 200):
                self.effective_proxy_list.append(str(proxy))           
        except : pass    
        
    def check_proxy(self,time_out,association_number):
        '''检查代理IP是否有效，有效则将其放进proxies.txt中'''  
        proxies_num = 0 
        read_file   = open(self.proxy_file1, 'r', encoding = 'utf-8')
        proxy_get   = read_file.readlines()
        read_file.close 
        if (len(proxy_get) != 0): #代理文件中有代理
            process_list = []
            pool = gevent.pool.Pool(association_number)         
            for ip in proxy_get:
                proxy = eval(ip.rstrip('\n')) #去除\n并转为字典
                process_list.append(pool.spawn(self.check, time_out,proxy))
            gevent.joinall(process_list)
            if (len(self.effective_proxy_list)): #有效代理IP列表不为空
                wirte_file = open(self.proxy_file2, 'w+', encoding = 'utf-8')
                for value in self.effective_proxy_list:
                    wirte_file.writelines(value + '\n') 
                wirte_file.close()  #获取到有效代理!
                proxies_num= len(self.effective_proxy_list)
            else : pass
        else : pass
        
        return proxies_num    
    
    def get_headers_and_proxy_list(self,proxy_setting,time_out,association_number):
        '''proxy_list列表：[value,value,...]'''
        proxy_list,process_list = [],[] #获取设置好的代理列表
        if (proxy_setting == 'off'): #不使用代理
            proxy_list.append(urllib.request.ProxyHandler(proxies = None))
        elif (proxy_setting == 'open'): #使用代理
            try :          
                read_file = open(self.proxy_file2,'r',encoding = 'utf-8')
                proxy_get = read_file.readlines()
                read_file.close
                def check(time_out,proxy): 
                    try :
                        response = requests.get(url='http://www.baidu.com',headers={'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'},timeout = time_out,proxies = proxy)
                        if (response.status_code == 200):
                            proxy_list.append(urllib.request.ProxyHandler(proxies = proxy))  
                    except : pass

                pool= gevent.pool.Pool(association_number)  
                for value in proxy_get:
                    process_list.append(pool.spawn(check, time_out,eval(value.rstrip('\n'))))
                gevent.joinall(process_list)                     
                if (len(proxy_list) == 0): #proxy_list没有获取到有效代理IP时，不设置代理
                    proxy_list.append(urllib.request.ProxyHandler(proxies = None))
            except : #打开文件失败时，也不设置代理
                proxy_list.append(urllib.request.ProxyHandler(proxies = None))

        else : pass

        return self.headers_list,proxy_list 
    

    def md5_get(self,value):
        '''得到value的MD5值'''
        x = hashlib.md5()
        x.update(value)
        md5_value = x.hexdigest()
        return md5_value    

    def file_link_dict_get(self,url,time_out,proxy_list):
        link_list = []
        request   = urllib.request.Request(url = url,headers = random.choice(self.headers_list)) 
        opener    = urllib.request.build_opener(random.choice(proxy_list)) #是否使用代理
        urllib.request.install_opener(opener)
        try :
            response = urllib.request.urlopen(request,timeout = time_out) #超时初始设置3 
            text     = response.read().decode("utf-8") 
            file_suffix_list = ['.png', '.ico', '.gif','.svg', '.jpeg','.js','.css','.xml','.txt'] #文件后缀名列表
            tags = ['a', 'A', 'link', 'script', 'area', 'iframe', 'form']  # img
            tos = ['href', 'src', 'action']
            for tag in tags:
                for to in tos:
                    link1 = re.findall(r'<%s.*?%s="(.*?)"' % (tag, to), text) #获取=""中的链接
                    link2 = re.findall(r'<%s.*?%s=\'(.*?)\'' % (tag, to), text) #获取=''中的链接
                    if (len(link1) != 0):
                        for i in link1:
                            link_list.append(i) 
                    else : pass
                    if (len(link2) != 0):
                        for i in link2:
                            if (i not in link_list):
                                link_list.append(i) #去重后合并链接   
                    else : pass
            if (len(link_list) != 0): #获取到了网页中的链接                
                for i in range(len(link_list)):
                    for j in range(len(file_suffix_list)):
                        if (file_suffix_list[j] in link_list[i][-4:]): #获取的链接中存在特定的后缀名
                            if re.findall(r':', link_list[i]): #查找链接中是否有“:”,如果存在可能是其他网址的链接，舍去
                                pass
                            else:
                                self.file_link_dict[url].append(url + link_list[i].replace('//' + url.replace('http://',''),'')) #去掉//url/...格式链接
                            if (re.findall(url,link_list[i])): #链接中自带有url就添加
                                self.file_link_dict[url].append(link_list[i])
                            else : pass
                    if ((i == len(link_list) - 1) and (len(self.file_link_dict[url]) == 0)): #没有在网页中匹配到指定文件后缀名的的文件链接
                        self.file_link_dict[url].append('Not_found')
                    else : pass
            else : #没有获取到网页中的链接
                self.file_link_dict[url].append('Not_found')
        except : #url请求失败
            self.file_link_dict[url].append('Not_found')

    def file_md5_list_get(self,url,time_out,proxy_list,cms_name,file_link):
        if (file_link != 'Not_found'):
            request = urllib.request.Request(url = file_link,headers = random.choice(self.headers_list)) 
            opener  = urllib.request.build_opener(random.choice(proxy_list)) #是否使用代理
            urllib.request.install_opener(opener)
            try :
                response  = urllib.request.urlopen(request,timeout = time_out) #超时初始设置3 
                read      = response.read()
                md5_value = self.md5_get(value = read)
                self.file_md5_list.append([cms_name,file_link.replace(url,''),md5_value,'md5'])
            except : #url请求失败
                self.file_md5_list.append(['Not_found'])                    
        else: 
            self.file_md5_list.append(['Not_found'])                               

    def get_file_md5_list(self,domain_list,cms_name_list,time_out,proxy_list,association_number):
        '''#获取file_md5_list列表[['cms_name','path','match_pattern','keyword'],...]'''
        process_list = []
        pool= gevent.pool.Pool(association_number)  
        for domain in domain_list:  
            self.file_link_dict.update({'http://' + domain:[]}) #初始化self.file_link_dict字典{'url':[],...}
            process_list.append(pool.spawn(self.file_link_dict_get,'http://' + domain,time_out,proxy_list))
        gevent.joinall(process_list)
        process_list = [] #清空协程池
        for i in range(len(domain_list)):
            for file_link in self.file_link_dict['http://' + domain_list[i]]:
                process_list.append(pool.spawn(self.file_md5_list_get,'http://' + domain_list[i],time_out,proxy_list,cms_name_list[i],file_link))
        gevent.joinall(process_list)

        return self.file_md5_list 

    def add_file_md5_to_cms_database(self,file_md5_list):
        new_file_md5_list = []
        cms_finger_name_get,cms_finger_path_get,cms_finger_match_pattern_get,cms_finger_options_get = [],[],[],[]
        workbook  = openpyxl.load_workbook((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'cms.xlsx')
        worksheet = workbook.active
        for cell1,cell2,cell3,cell4 in zip(worksheet['A'][1:],worksheet['B'][1:],worksheet['C'][1:],worksheet['D'][1:]):
            cms_finger_name_get.append(cell1.value)
            cms_finger_path_get.append(cell2.value)
            cms_finger_match_pattern_get.append(cell3.value)
            cms_finger_options_get.append(cell4.value)
        workbook.close()
        for value in file_md5_list:
            if ((value != ['Not_found']) and (value[1] in cms_finger_path_get) and (value[2] not in cms_finger_match_pattern_get)): #路径在库，md5值不在库
                new_file_md5_list.append(value)
            elif ((value != ['Not_found']) and (value[1] not in cms_finger_path_get)):
                new_file_md5_list.append(value)
            else : pass

        if (len(new_file_md5_list) != 0): #要添进指纹库的指纹列表不为空，进行指纹添加
            workbook  = openpyxl.load_workbook((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'cms.xlsx')
            worksheet = workbook.active #获取表的cms表单(默认最活跃的即是第一个表)
            for i in range(len(new_file_md5_list)):
                worksheet.append([new_file_md5_list[i][0], new_file_md5_list[i][1],new_file_md5_list[i][2],new_file_md5_list[i][3],0])
            workbook.save((os.getcwd() + '/Fingerprint_database/').replace('\\','/') + 'cms.xlsx')
            workbook.close
        else : pass
        
    def output_save(self,datas,save_type): 
        data_list = []
        for value in datas: #Python列表对象被引用时，函数内的引用对象的改变，会导致函数外的对象也会改变
            if (value not in data_list):
                data_list.append(value)
        data_list.insert(0,['域名','网站标题','IP地址','CMS','Banner信息','操作系统','服务器','JS框架','开发语言'])
        now_time = time.strftime("%Y-%m-%d-%H_%M_%S",time.localtime(time.time())) #获取当前时间
        if (save_type == 'Csv'):
            file_name = (os.getcwd().replace('Cms_get','') + '/Result_output/').replace('\\','/')  + now_time + r"_result.csv"
            file_csv = codecs.open(file_name,'w',encoding = 'utf-8') #用Excel打开文件会出现乱码情况，是正常的，更改Excel的编码格式就好了
            writer = csv.writer(file_csv, delimiter=' ', quotechar=' ', quoting=csv.QUOTE_MINIMAL)
            for value in data_list:
                writer.writerow(value)   
        elif (save_type == 'Excel'):
            file_name = (os.getcwd().replace('Cms_get','') + '/Result_output/').replace('\\','/')  + now_time + r"_result.xls"      
            file = xlwt.Workbook()
            result = file.add_sheet(u'result',cell_overwrite_ok = True) #创建sheet
            i = 0
            for value in data_list:
                for j in range(len(value)):
                    result.write(i,j,value[j])
                i = i + 1    
            file.save(file_name) 
        else :
            file_name = (os.getcwd().replace('Cms_get','') + '/Result_output/').replace('\\','/')  + now_time + r"_result.txt"    
            file = open(file_name,'w',encoding = 'utf-8')
            for i in range(len(data_list)):
                data = str(data_list[i]).replace('[','').replace(']','')#去除[],这两行按数据不同，可以选择
                data = data.replace("'",'').replace(',','--||--') +'\n'   #去除单引号，替换逗号为“--||--”，每行末尾追加换行符
                file.write(data)
            file.close()      
        
    
        
